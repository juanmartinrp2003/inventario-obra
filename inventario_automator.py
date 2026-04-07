"""
╔══════════════════════════════════════════════════════════════════╗
║   AUTOMATIZADOR DE INVENTARIO — MULTI-PROYECTO                  ║
║   Distribución de inventario por rubros según órdenes de pedido ║
╚══════════════════════════════════════════════════════════════════╝
Uso: streamlit run inventario_automator.py
"""
 
import re
import io
import json
import copy
import shutil
import unicodedata
from pathlib import Path
from datetime import datetime, date
from difflib import SequenceMatcher
from typing import Optional
 
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# ─────────────────────────────────────────────────────────────────────────────
# PROJECT PERSISTENCE  (projects.json en el mismo directorio)
# ─────────────────────────────────────────────────────────────────────────────
PROJECTS_FILE = Path(__file__).parent / "projects.json"
 
def load_projects() -> list[str]:
    """Load project list from projects.json, creating it if absent."""
    if PROJECTS_FILE.exists():
        try:
            data = json.loads(PROJECTS_FILE.read_text(encoding="utf-8"))
            return data.get("projects", [])
        except Exception:
            pass
    return []
 
def save_projects(projects: list[str]) -> None:
    """Persist project list to projects.json."""
    PROJECTS_FILE.write_text(
        json.dumps({"projects": projects}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def get_project_mappings(project: str) -> dict:
    """Return manual mappings {inv_norm: order_norm} for the given project."""
    return st.session_state.mappings.get(project, {})


def save_project_mappings(project: str, mappings: dict) -> None:
    """Persist manual mappings for the given project in session_state."""
    st.session_state.mappings[project] = mappings
 
# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
INV_SHEET = "Inv."
HEADER_ROW = 4           # Row with rubro codes (G4 onwards)
MAT_START_ROW = 6        # First material row
MAT_END_ROW = 120        # Fallback — auto-detected at runtime from the sheet
TOTAL_ROW = 121
PET_START_ROW = 122
PET_END_ROW = 127
TOTAL_PET_ROW = 128
TOTAL_INV_ROW = 129
COL_MATERIAL = 1         # A
COL_ONHAND = 2           # B
COL_AVGCOST = 4          # D
COL_ASSETVAL = 5         # E
RUBRO_COL_START = 7      # G
FUZZY_THRESHOLD = 0.72   # Minimum similarity for "dubious" match
 
# ─────────────────────────────────────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
 
def _clean_text(s: str) -> str:
    """Apply accent stripping, symbol normalization, and spacing cleanup.
    Does NOT remove parenthetical content — call this on already-stripped text."""
    # Strip accents/diacritics (á→a, é→e, ó→o, ñ→n, etc.)
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    # Normalize number symbols ONLY when followed by a digit, to avoid
    # matching "no" in regular words like "hormigono", "rodano", etc.
    #   N°8, No.8, No8  → n.8
    #   #8              → n.8
    s = re.sub(r'\b[Nn][°o]\.?\s*(?=\d)', 'n.', s)
    s = re.sub(r'#\s*(?=\d)', 'n.', s)
    # Normalize spaces after periods before digits: "N. 18" → "n.18"
    s = re.sub(r'\.\s+(\d)', r'.\1', s)
    # Collapse multiple spaces
    s = re.sub(r'\s+', ' ', s).strip()
    return s.lower()
 
 
def normalize(text: str) -> str:
    """Normalize a material name for matching.
 
    Key behaviours:
    - Strips accents, lowercases, normalises N°/No./#
    - Removes parenthetical content — BUT if the FIRST parenthetical block is
      LONGER than the text before it, that block is used instead (it is the
      full description while the base text is just an abbreviation).
      e.g. 'Porcelanato Rodano Chalk Estruc (Porcelanato Rodano Chalk
            Estructurado 60x60 1.8m)'
           → 'porcelanato rodano chalk estructurado 60x60 1.8m'
    - After choosing base vs paren content, removes any remaining
      parenthetical content iteratively.
    """
    if not text:
        return ""
    s = str(text).strip()
 
    # Split base text (before first paren) and first paren content
    base_text = re.sub(r'\s*\(.*', '', s).strip()
    first_paren_m = re.search(r'\(([^()]+)\)', s)
 
    if first_paren_m and len(first_paren_m.group(1).strip()) > len(base_text):
        # The paren content is the full name; use it
        s = first_paren_m.group(1).strip()
    else:
        # Normal case: remove ALL parenthetical content iteratively
        prev = None
        while prev != s:
            prev = s
            s = re.sub(r'\s*\([^()]*\)\s*', ' ', s).strip()
 
    return _clean_text(s)
 
 
def extract_rubro_code(rubro_cell: str) -> str:
    """Extract numeric code from 'XX.XX.XX Description' string.
    Returns '' if no XX.XX.XX pattern is found."""
    if not rubro_cell:
        return ""
    m = re.match(r'^\s*(\d+\.\d+\.\d+)', str(rubro_cell).strip())
    return m.group(1) if m else ""
 
 
def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()
 
 
def token_sort_ratio(a: str, b: str) -> float:
    """Compare strings after sorting their tokens alphabetically.
    Handles cases where word order differs between inventory and orders."""
    a_sorted = ' '.join(sorted(a.split()))
    b_sorted = ' '.join(sorted(b.split()))
    return SequenceMatcher(None, a_sorted, b_sorted).ratio()
 
 
def prefix_match_score(a: str, b: str) -> float:
    """Returns a high score when one string is a prefix of the other.
    Handles: 'bisagra' vs 'bisagras para puerta', 'zapapico' vs 'zapapico herrago 5 lbs',
             'cabo 1/2' vs 'cabo', 'maxicril - 70 gl' vs 'maxicril'.
    Score = 0.80 + 0.20 * coverage, so always above FUZZY_THRESHOLD.
    Requires the shorter string to be at least 4 characters.
    """
    if not a or not b:
        return 0.0
    short, long = (a, b) if len(a) <= len(b) else (b, a)
    if len(short) < 4:
        return 0.0
    if long.startswith(short):
        coverage = len(short) / len(long)
        return 0.80 + 0.20 * coverage
    return 0.0
 
 
def best_similarity(a: str, b: str) -> float:
    """Returns the best score across direct, token-sort, and prefix matching."""
    return max(similarity(a, b), token_sort_ratio(a, b), prefix_match_score(a, b))
 
 
def parse_date(value) -> Optional[date]:
    """Convert various date representations to a date object."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    # Try common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d",
                "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None
 
 
# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — PARSE INVENTORY
# ─────────────────────────────────────────────────────────────────────────────
 
def _detect_row_structure(ws) -> dict:
    """
    Scan the Inv. sheet starting from MAT_START_ROW and locate the key
    sentinel rows by matching the text in column A (case-insensitive):
 
        "TOTAL"                    → total_row      (end of regular materials)
        contains "PETREO"          → total_pet_row  (end of pétreos block)
        contains "TOTAL INVENTARIO"→ total_inv_row
 
    Returns a dict with the detected row numbers, falling back to the
    module-level constants if a sentinel is not found.
    """
    total_row = TOTAL_ROW
    total_pet_row = TOTAL_PET_ROW
    total_inv_row = TOTAL_INV_ROW
 
    for row in range(MAT_START_ROW, ws.max_row + 1):
        raw = ws.cell(row, COL_MATERIAL).value
        if raw is None:
            continue
        label = str(raw).strip().upper()
 
        if label == "TOTAL" and total_row == TOTAL_ROW:
            total_row = row
        elif "PETREO" in label and total_pet_row == TOTAL_PET_ROW:
            total_pet_row = row
        elif "TOTAL INVENTARIO" in label and total_inv_row == TOTAL_INV_ROW:
            total_inv_row = row
            break  # Nothing useful below this
 
    mat_end_row  = total_row - 1
    pet_start_row = total_row + 1
    pet_end_row   = total_pet_row - 1
 
    return {
        "mat_end_row":    mat_end_row,
        "total_row":      total_row,
        "pet_start_row":  pet_start_row,
        "pet_end_row":    pet_end_row,
        "total_pet_row":  total_pet_row,
        "total_inv_row":  total_inv_row,
    }
 
 
def _read_material_rows(ws, start: int, end: int) -> list:
    """Read material rows from `start` to `end` (inclusive), skipping on_hand=0."""
    items = []
    for row in range(start, end + 1):
        name     = ws.cell(row, COL_MATERIAL).value
        on_hand  = ws.cell(row, COL_ONHAND).value
        avg_cost = ws.cell(row, COL_AVGCOST).value
        asset_val = ws.cell(row, COL_ASSETVAL).value
 
        if name is None or on_hand is None:
            continue
        try:
            on_hand_f = float(on_hand)
        except (TypeError, ValueError):
            continue
        if on_hand_f == 0:
            continue
 
        items.append({
            "row":        row,
            "name":       str(name).strip(),
            "norm_name":  normalize(str(name)),
            "on_hand":    on_hand_f,
            "avg_cost":   float(avg_cost) if avg_cost else 0.0,
            "asset_value": float(asset_val) if asset_val else 0.0,
        })
    return items
 
 
def parse_inventory(wb: openpyxl.Workbook) -> dict:
    """
    Returns:
        {
          'rubros':        {code: col_index},
          'materials':     [...],   # regular materials with on_hand > 0
          'petreos':       [...],   # pétreos/services with on_hand > 0
          'rubro_last_col': int,
          'structure':     {mat_end_row, total_row, pet_start_row, ...}
        }
    """
    ws = wb[INV_SHEET]
 
    # ── Auto-detect sentinel rows ─────────────────────────────────────────────
    structure = _detect_row_structure(ws)
 
    # ── Read rubro headers from row 4, cols G onwards ─────────────────────────
    rubros = {}
    rubro_last_col = RUBRO_COL_START
    for col in range(RUBRO_COL_START, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, col).value
        if val is not None:
            code = extract_rubro_code(str(val))
            if code:
                rubros[code] = col
                rubro_last_col = col
 
    # ── Read materials and pétreos ────────────────────────────────────────────
    materials = _read_material_rows(ws, MAT_START_ROW, structure["mat_end_row"])
    petreos   = _read_material_rows(ws, structure["pet_start_row"], structure["pet_end_row"])
 
    return {
        "rubros":         rubros,
        "materials":      materials,
        "petreos":        petreos,
        "rubro_last_col": rubro_last_col,
        "structure":      structure,
    }
 
 
# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — PARSE ORDERS
# ─────────────────────────────────────────────────────────────────────────────
 
def _find_date_in_sheet(ws) -> Optional[date]:
    """Scan rows 1-12 for a cell containing 'FECHA' and return the date nearby."""
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=False):
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == "FECHA":
                # Look right in the same row for a non-None, non-label value
                for sibling in ws[cell.row]:
                    if sibling.column > cell.column and sibling.value is not None:
                        d = parse_date(sibling.value)
                        if d:
                            return d
    return None
 
 
def _find_header_row(ws):
    """
    Find the row that contains 'CANTIDAD' (the column header row).
    Returns (row_number, col_map) where col_map maps field→col_index (1-based).
    """
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
        cells = {str(c.value).strip().upper(): c.column
                 for c in row if c.value is not None}
        if "CANTIDAD" in cells:
            return row[0].row, cells
    return None, {}
 
 
def parse_orders(wb_orders: openpyxl.Workbook, cutoff: date) -> list:
    """
    Returns list of order_lines:
        {pedido, fecha, qty, unit, material, norm_material, rubro_code, rubro_full, estado}
    Only includes sheets with fecha <= cutoff.
    """
    lines = []
    for sheet_name in wb_orders.sheetnames:
        # Skip non-pedido sheets
        if not re.search(r'pedido', sheet_name, re.IGNORECASE):
            continue
 
        ws = wb_orders[sheet_name]
 
        # Extract order number from sheet name
        num_match = re.search(r'(\d+)', sheet_name)
        pedido_num = int(num_match.group(1)) if num_match else 0
 
        # Get date
        fecha = _find_date_in_sheet(ws)
        if fecha is None:
            # Skip if no date found
            continue
        if fecha > cutoff:
            continue  # Outside cutoff
 
        # Find header row
        hdr_row, col_map = _find_header_row(ws)
        if hdr_row is None:
            continue
 
        # Determine column indices (prefer common names)
        def get_col(*names):
            for n in names:
                if n in col_map:
                    return col_map[n]
            return None
 
        col_qty = get_col("CANTIDAD")
        col_mat = get_col("MATERIAL")
        col_rub = get_col("RUBRO")
        col_cap = get_col("CAPITULO")   # Some pedidos store the code here
        col_est = get_col("ESTADO")
        col_und = get_col("UND.", "UND", "UNIDAD")
 
        if not (col_qty and col_mat and (col_rub or col_cap)):
            continue
 
        # Read detail rows (from hdr_row+1 until empty CANTIDAD)
        for row_idx in range(hdr_row + 1, ws.max_row + 1):
            qty_val = ws.cell(row_idx, col_qty).value
            mat_val = ws.cell(row_idx, col_mat).value
            rub_val = ws.cell(row_idx, col_rub).value if col_rub else None
            cap_val = ws.cell(row_idx, col_cap).value if col_cap else None
 
            if qty_val is None and mat_val is None:
                break  # End of data section
 
            try:
                qty = float(qty_val) if qty_val is not None else 0.0
            except (TypeError, ValueError):
                continue
            if qty <= 0:
                continue
            if mat_val is None:
                continue
 
            material_str = str(mat_val).strip()
            rubro_str = str(rub_val).strip() if rub_val else ""
            cap_str   = str(cap_val).strip() if cap_val else ""
 
            # Try RUBRO column first; fall back to CAPITULO when RUBRO has no
            # numeric code (e.g. Pedido 49 stores "08.01.09" in CAPITULO).
            rubro_code = extract_rubro_code(rubro_str) or extract_rubro_code(cap_str)
            # Use whichever column has the full descriptive text
            rubro_full = rubro_str if rubro_str else cap_str
 
            estado = str(ws.cell(row_idx, col_est).value).strip() if col_est else "N/A"
            unit = str(ws.cell(row_idx, col_und).value).strip() if col_und else ""
 
            lines.append({
                "pedido": pedido_num,
                "pedido_name": sheet_name,
                "fecha": fecha,
                "qty": qty,
                "unit": unit,
                "material": material_str,
                "norm_material": normalize(material_str),
                "rubro_code": rubro_code,
                "rubro_full": rubro_full,
                "estado": estado,
            })
 
    return lines
 
 
# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — MATCH MATERIALS
# ─────────────────────────────────────────────────────────────────────────────
 
def match_materials(inv_materials: list, order_lines: list,
                    manual_mappings: dict = None) -> tuple:
    """
    Returns:
        matched_lines: same list but each line has 'inv_row' and 'match_type'
                       ('exact', 'fuzzy', 'manual', 'none')
        list_a: unmatched inventory materials (on_hand>0, no orders)
        list_b: dubious matches
    """
    # Build reverse lookup order_norm → inv_norm from manual mappings
    order_to_inv: dict = {}
    if manual_mappings:
        for inv_norm, order_norm in manual_mappings.items():
            order_to_inv[order_norm] = inv_norm

    # Build a lookup set of normalized inventory names → material record
    inv_by_norm = {m["norm_name"]: m for m in inv_materials}
    inv_names = list(inv_by_norm.keys())
 
    matched_lines = []
    list_b = []
 
    for line in order_lines:
        norm = line["norm_material"]
        match_type = "none"
        inv_row = None
        inv_name_matched = None
 
        # Manual mapping (highest priority — overrides exact and fuzzy)
        if norm in order_to_inv:
            mapped_inv_norm = order_to_inv[norm]
            if mapped_inv_norm in inv_by_norm:
                inv_row = inv_by_norm[mapped_inv_norm]["row"]
                inv_name_matched = inv_by_norm[mapped_inv_norm]["name"]
                match_type = "manual"

        # Exact match
        if match_type == "none" and norm in inv_by_norm:
            inv_row = inv_by_norm[norm]["row"]
            inv_name_matched = inv_by_norm[norm]["name"]
            match_type = "exact"
        elif match_type == "none":
            # Fuzzy match (uses token-sort as well to handle different word orders)
            best_score = 0.0
            best_name = None
            for inv_norm in inv_names:
                score = best_similarity(norm, inv_norm)
                if score > best_score:
                    best_score = score
                    best_name = inv_norm
            if best_score >= FUZZY_THRESHOLD:
                inv_row = inv_by_norm[best_name]["row"]
                inv_name_matched = inv_by_norm[best_name]["name"]
                match_type = "fuzzy"
                list_b.append({
                    "name_inv": inv_name_matched,
                    "name_order": line["material"],
                    "norm_inv": best_name,
                    "norm_order": norm,
                    "pedido": line["pedido_name"],
                    "rubro": line["rubro_full"],
                    "score": round(best_score, 3),
                })
 
        matched_line = dict(line)
        matched_line["inv_row"] = inv_row
        matched_line["match_type"] = match_type
        matched_line["inv_name_matched"] = inv_name_matched
        matched_lines.append(matched_line)
 
    # List A: materials with on_hand > 0 that have NO matched lines
    rows_with_matches = {l["inv_row"] for l in matched_lines if l["inv_row"]}
    list_a = [m for m in inv_materials if m["row"] not in rows_with_matches]
 
    return matched_lines, list_a, list_b
 
 
# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — DISTRIBUTION ENGINE
# ─────────────────────────────────────────────────────────────────────────────
 
def distribute(inv_materials: list, matched_lines: list, rubros: dict) -> tuple:
    """
    Apply the LIFO distribution logic (most recent order first).
 
    Returns:
        allocations:      {inv_row: {rubro_code: value_assigned}}
        list_c:           lines from PENDING orders that were used
        unmatched_rubros: rubro codes from orders not found in inventory header
        remainder_alerts: materials where on_hand > total ordered qty
                          (remainder auto-assigned to most recent rubro)
    """
    from collections import defaultdict
    lines_by_row = defaultdict(list)
    for l in matched_lines:
        if l["inv_row"] is not None:
            lines_by_row[l["inv_row"]].append(l)
 
    allocations = {}
    list_c = []
    unmatched_rubros = set()
    remainder_alerts = []
 
    inv_by_row = {m["row"]: m for m in inv_materials}
 
    for inv_row, lines in lines_by_row.items():
        mat = inv_by_row.get(inv_row)
        if not mat:
            continue
 
        on_hand = mat["on_hand"]
        avg_cost = mat["avg_cost"]
 
        # Sort: most recent date first, then by pedido desc (stable)
        lines_sorted = sorted(lines, key=lambda l: (l["fecha"], l["pedido"]), reverse=True)
 
        remaining = on_hand
        row_alloc = defaultdict(float)  # rubro_code → qty
 
        for line in lines_sorted:
            if remaining <= 0:
                break
 
            rubro_code = line["rubro_code"]
            if rubro_code not in rubros:
                # Rubro not found in inventory headers — record and SKIP.
                # Do NOT consume remaining; otherwise the qty is silently lost.
                unmatched_rubros.add((rubro_code, line["rubro_full"]))
                continue
 
            take = min(line["qty"], remaining)
            row_alloc[rubro_code] += take
            remaining -= take
 
            # Track List C (pending lines used)
            estado_lower = line["estado"].lower()
            if any(w in estado_lower for w in ("pendiente", "pending")):
                list_c.append({
                    "material": mat["name"],
                    "pedido": line["pedido_name"],
                    "fecha": line["fecha"].strftime("%d/%m/%Y"),
                    "qty_used": round(take, 4),
                    "rubro": line["rubro_full"],
                    "estado": line["estado"],
                })
 
        # ── Remainder: on_hand exceeds total ordered qty ──────────────────────
        if remaining > 1e-9:  # float tolerance
            # Assign remainder to the most recent order's rubro
            most_recent = lines_sorted[0]
            fallback_rubro = most_recent["rubro_code"]
            row_alloc[fallback_rubro] += remaining
            remainder_alerts.append({
                "material":         mat["name"],
                "on_hand":          on_hand,
                "cubierto_ordenes": round(on_hand - remaining, 4),
                "sobrante":         round(remaining, 4),
                "rubro_asignado":   most_recent["rubro_full"],
                "pedido_referencia": most_recent["pedido_name"],
                "fecha_referencia": most_recent["fecha"].strftime("%d/%m/%Y"),
            })
 
        # Convert qty → value
        allocations[inv_row] = {
            code: round(qty * avg_cost, 2)
            for code, qty in row_alloc.items()
        }

        # ── Rounding correction ───────────────────────────────────────────────
        # Compare asset_value against the sum that will *actually be written*
        # (only rubros that exist in the inventory header). This handles both:
        #   a) pure per-rubro rounding drift
        #   b) cases where the remainder was assigned to an invalid rubro
        # If the gap is > 0 and < 1, absorb it into the most recent valid rubro.
        asset_value = mat["asset_value"]
        written_sum = sum(v for code, v in allocations[inv_row].items()
                          if code in rubros)
        gap = asset_value - written_sum          # do NOT round before comparing
        if 1e-9 < gap < 1:
            # Walk LIFO order to find the most recent rubro that is valid
            fallback = None
            for line in lines_sorted:
                rc = line["rubro_code"]
                if rc in rubros:
                    if rc not in allocations[inv_row]:
                        allocations[inv_row][rc] = 0.0
                    fallback = rc
                    break
            if fallback is not None:
                allocations[inv_row][fallback] = round(
                    allocations[inv_row][fallback] + gap, 2
                )

    return allocations, list_c, list(unmatched_rubros), remainder_alerts
 
 
# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — WRITE RESULTS TO WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────
 
def write_to_workbook(wb: openpyxl.Workbook, allocations: dict, rubros: dict,
                      rubro_last_col: int, structure: dict) -> openpyxl.Workbook:
    """
    Writes allocation values into Inv. sheet and inserts summary formulas.
    Uses `structure` (from _detect_row_structure) for all row boundaries —
    so the function works correctly regardless of how many material rows exist.
    """
    ws = wb[INV_SHEET]
 
    mat_end_row   = structure["mat_end_row"]
    total_row     = structure["total_row"]
    pet_start_row = structure["pet_start_row"]
    pet_end_row   = structure["pet_end_row"]
    total_pet_row = structure["total_pet_row"]
    total_inv_row = structure["total_inv_row"]
 
    # Clear existing values in rubro columns for all relevant rows
    for col in rubros.values():
        for row in range(MAT_START_ROW, total_inv_row + 1):
            ws.cell(row, col).value = None
 
    # Write allocation values
    for inv_row, rubro_vals in allocations.items():
        for rubro_code, value in rubro_vals.items():
            if rubro_code in rubros:
                col = rubros[rubro_code]
                current = ws.cell(inv_row, col).value or 0
                ws.cell(inv_row, col).value = round(float(current) + value, 2)
 
    # Write summary formulas for each rubro column
    for code, col in rubros.items():
        col_letter = get_column_letter(col)
        ws.cell(total_row,     col).value = f"=SUM({col_letter}{MAT_START_ROW}:{col_letter}{mat_end_row})"
        ws.cell(total_pet_row, col).value = f"=SUM({col_letter}{pet_start_row}:{col_letter}{pet_end_row})"
        ws.cell(total_inv_row, col).value = f"={col_letter}{total_row}+{col_letter}{total_pet_row}"
 
    # Write Verificación column formulas
    ver_col = rubro_last_col + 1
    vcl      = get_column_letter(ver_col)
    rub_s    = get_column_letter(RUBRO_COL_START)
    rub_e    = get_column_letter(rubro_last_col)
    e_col    = get_column_letter(COL_ASSETVAL)
 
    for row in range(MAT_START_ROW, mat_end_row + 1):
        ws.cell(row, ver_col).value = f"=SUM({rub_s}{row}:{rub_e}{row})-{e_col}{row}"
    ws.cell(total_row,     ver_col).value = f"=SUM({vcl}{MAT_START_ROW}:{vcl}{mat_end_row})"
    ws.cell(total_pet_row, ver_col).value = f"=SUM({vcl}{pet_start_row}:{vcl}{pet_end_row})"
    ws.cell(total_inv_row, ver_col).value = f"={vcl}{total_row}+{vcl}{total_pet_row}"
 
    return wb
 
 
# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT UI — MULTI-PROYECTO
# ─────────────────────────────────────────────────────────────────────────────
 
st.set_page_config(
    page_title="Inventario de Obra — Multi-Proyecto",
    page_icon="🏗️",
    layout="wide",
)
 
# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main { background: #f8f9fb; }
    .block-container { padding-top: 2rem; }
    .stMetric { background: white; border-radius: 10px; padding: 1rem;
                border: 1px solid #e0e0e0; box-shadow: 0 1px 4px rgba(0,0,0,0.06); }
    .section-header {
        background: linear-gradient(90deg, #1a3a5c 0%, #2d6099 100%);
        color: white; padding: 0.6rem 1.2rem; border-radius: 8px;
        font-weight: bold; margin: 1.2rem 0 0.6rem 0;
    }
    .project-badge {
        display:inline-block; background:#e8f4fd; color:#1a3a5c;
        border:1px solid #2d6099; border-radius:20px;
        padding:4px 14px; font-weight:700; font-size:1rem; margin-bottom:8px;
    }
</style>
""", unsafe_allow_html=True)
 
st.title("🏗️ Automatizador de Inventario de Obra")
st.markdown("Distribución automática de inventario por rubros según órdenes de pedido.")
 
# ── Session state init ────────────────────────────────────────────────────────
if "projects" not in st.session_state:
    st.session_state.projects = load_projects()
if "active_project" not in st.session_state:
    st.session_state.active_project = None
if "mappings" not in st.session_state:
    st.session_state.mappings = {}
if "auto_run" not in st.session_state:
    st.session_state.auto_run = False
 
# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("🗂️ Proyectos")
 
    projects = st.session_state.projects
 
    # ── Select existing project ───────────────────────────────────────────────
    if projects:
        selected = st.selectbox(
            "Seleccionar proyecto",
            options=projects,
            index=0,
        )
        st.session_state.active_project = selected
    else:
        st.info("No hay proyectos creados. Crea el primero abajo.")
        st.session_state.active_project = None
 
    # ── Create new project ────────────────────────────────────────────────────
    with st.expander("➕ Crear nuevo proyecto"):
        new_name = st.text_input(
            "Nombre del proyecto",
            placeholder="Ej: Residencia MH, Edificio Centro, Villa Norte…",
        )
        if st.button("Crear proyecto", type="primary"):
            name = new_name.strip()
            if not name:
                st.error("El nombre no puede estar vacío.")
            elif name in st.session_state.projects:
                st.warning(f'El proyecto "{name}" ya existe.')
            else:
                st.session_state.projects.append(name)
                save_projects(st.session_state.projects)
                st.session_state.active_project = name
                st.success(f'✅ Proyecto "{name}" creado.')
                st.rerun()
 
    # ── Delete project ────────────────────────────────────────────────────────
    if projects:
        with st.expander("🗑️ Eliminar proyecto"):
            del_name = st.selectbox("Proyecto a eliminar", options=projects, key="del_sel")
            if st.button("Eliminar", type="secondary"):
                st.session_state.projects.remove(del_name)
                save_projects(st.session_state.projects)
                if st.session_state.active_project == del_name:
                    st.session_state.active_project = (
                        st.session_state.projects[0] if st.session_state.projects else None
                    )
                st.rerun()
 
    st.divider()
 
    # ── Only show file inputs if a project is active ──────────────────────────
    active = st.session_state.active_project
    if active:
        st.subheader(f"📂 Archivos — {active}")
 
        cutoff_date = st.date_input(
            "📅 Fecha de corte",
            value=date(2025, 12, 31),
            help="Solo se usan las órdenes con fecha ≤ a esta fecha.",
        )
 
        inv_file = st.file_uploader(
            "Reporte CONSOLIDADO",
            type=["xlsx"],
            help="Archivo con la hoja 'Inv.'",
            key=f"inv_{active}",
        )
        orders_file = st.file_uploader(
            "Órdenes de Pedido",
            type=["xlsx"],
            help="Archivo con hojas Pedido01, Pedido02…",
            key=f"ord_{active}",
        )

        # ── Mapeos manuales ───────────────────────────────────────────────────
        st.divider()
        st.caption("📎 Mapeos manuales")
        mappings_upload = st.file_uploader(
            "Cargar mapeos (JSON)",
            type=["json"],
            help="Sube el archivo JSON de mapeos guardado en una sesión anterior.",
            key=f"mapfile_{active}",
        )
        if mappings_upload:
            try:
                loaded = json.loads(mappings_upload.read().decode("utf-8"))
                save_project_mappings(active, loaded)
                st.success(f"✅ {len(loaded)} mapeos cargados.")
            except Exception as e:
                st.error(f"Error al cargar mapeos: {e}")

        current_maps = get_project_mappings(active)
        if current_maps:
            st.caption(f"🔗 {len(current_maps)} mapeos manuales activos")
            maps_json = json.dumps(current_maps, ensure_ascii=False, indent=2)
            st.download_button(
                "⬇ Descargar mapeos JSON",
                data=maps_json,
                file_name=f"mapeos_{re.sub(r'[^\\w-]', '_', active)}.json",
                mime="application/json",
                key="dl_maps_sidebar",
            )

        st.divider()
        st.caption(f"Umbral match fuzzy: **{FUZZY_THRESHOLD}**")
 
        run_btn = st.button(
            "▶ Ejecutar distribución",
            type="primary",
            disabled=(inv_file is None or orders_file is None),
        )
    else:
        inv_file = None
        orders_file = None
        run_btn = False
 
# ── Main area ─────────────────────────────────────────────────────────────────
active = st.session_state.active_project
 
if not active:
    st.info("⬅ Crea o selecciona un proyecto en el panel lateral para comenzar.")
    st.stop()
 
# Project badge
st.markdown(f'<div class="project-badge">📁 {active}</div>', unsafe_allow_html=True)
 
if inv_file is None or orders_file is None:
    st.markdown("""
    **Flujo del proceso:**
    1. Selecciona el **proyecto** en el panel lateral
    2. Sube el **Reporte CONSOLIDADO** (con hoja "Inv.")
    3. Sube el archivo de **Órdenes de Pedido** (Pedido01…PedidoN)
    4. Selecciona la **Fecha de corte**
    5. Haz clic en **▶ Ejecutar distribución**
    6. Revisa el reporte y descarga el archivo actualizado
    """)
    st.stop()
 
# ── Process ───────────────────────────────────────────────────────────────────
if run_btn or st.session_state.auto_run:
    st.session_state.auto_run = False
    with st.spinner("Procesando… esto puede tardar unos segundos."):
 
        # Load workbooks
        inv_bytes = inv_file.read()
        ord_bytes = orders_file.read()
 
        wb_inv = load_workbook(io.BytesIO(inv_bytes), data_only=True)
        wb_inv_write = load_workbook(io.BytesIO(inv_bytes))   # keeps formulas for writing
        wb_ord = load_workbook(io.BytesIO(ord_bytes), data_only=True)
 
        # Step 1 – Inventory
        inv_data = parse_inventory(wb_inv)
        rubros = inv_data["rubros"]
        materials = inv_data["materials"]
        petreos = inv_data["petreos"]
        all_materials = materials + petreos
        rubro_last_col = inv_data["rubro_last_col"]
 
        # Step 2 – Orders
        order_lines = parse_orders(wb_ord, cutoff_date)
        total_orders_sheets = sum(
            1 for s in wb_ord.sheetnames
            if re.search(r'pedido', s, re.IGNORECASE)
        )
        included_sheets = len({l["pedido_name"] for l in order_lines})
        excluded_sheets = total_orders_sheets - included_sheets
 
        # Step 3 – Matching
        manual_mappings = get_project_mappings(active)
        matched_lines, list_a, list_b = match_materials(all_materials, order_lines, manual_mappings)
        manual_match_count = sum(1 for l in matched_lines if l["match_type"] == "manual")
 
        # Step 4 – Distribution
        allocations, list_c, unmatched_rubros, remainder_alerts = distribute(
            all_materials, matched_lines, rubros
        )
 
        # Step 5 – Write
        wb_out = write_to_workbook(wb_inv_write, allocations, rubros,
                                   rubro_last_col, inv_data["structure"])
 
        # Serialize output
        out_buf = io.BytesIO()
        wb_out.save(out_buf)
        out_buf.seek(0)
        out_bytes = out_buf.read()
 
    st.success("✅ Distribución completada exitosamente.")
 
    # ── KPI Strip ─────────────────────────────────────────────────────────────
    col1, col2, col3, col4, col5, col6, col7 = st.columns(7)
    col1.metric("📦 Materiales con stock", len(all_materials))
    col2.metric("📋 Órdenes incluidas", included_sheets,
                delta=f"-{excluded_sheets} excluidas")
    col3.metric("🔗 Líneas de pedido", len(order_lines))
    col4.metric("⚠ Sin match (Lista A)", len(list_a))
    col5.metric("🔍 Match dudoso (Lista B)", len(list_b))
    col6.metric("🟡 Sobrante asignado", len(remainder_alerts))
    col7.metric("🔗 Mapeos aplicados", manual_match_count)
 
    # ── Download ──────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">💾 Archivo procesado</div>',
                unsafe_allow_html=True)
    safe_project = re.sub(r'[^\w\-]', '_', active)
    fname = f"{safe_project}_Distribuido_{cutoff_date.strftime('%d-%m-%Y')}.xlsx"
    st.download_button(
        label=f"⬇ Descargar CONSOLIDADO — {active}",
        data=out_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
 
    # ── Allocation summary table ───────────────────────────────────────────────
    st.markdown('<div class="section-header">📊 Distribución por material</div>',
                unsafe_allow_html=True)
 
    # Build summary dataframe
    mat_by_row = {m["row"]: m for m in all_materials}
    summary_rows = []
    for inv_row, rub_vals in allocations.items():
        mat = mat_by_row.get(inv_row, {})
        total_assigned = sum(rub_vals.values())
        asset_val = mat.get("asset_value", 0)
        diff = round(total_assigned - asset_val, 2)
        summary_rows.append({
            "Material": mat.get("name", "?"),
            "On Hand": mat.get("on_hand", 0),
            "Asset Value ($)": asset_val,
            "Asignado ($)": round(total_assigned, 2),
            "Diferencia ($)": diff,
            "Rubros": len(rub_vals),
        })
 
    if summary_rows:
        df_summary = pd.DataFrame(summary_rows)
        st.dataframe(
            df_summary.style.format({
                "Asset Value ($)": "${:,.2f}",
                "Asignado ($)": "${:,.2f}",
                "Diferencia ($)": "${:,.2f}",
            }).map(
                lambda v: "background-color: #ffe0e0" if isinstance(v, float) and abs(v) > 0.05 else "",
                subset=["Diferencia ($)"]
            ),
            use_container_width=True,
            height=350,
        )
 
    # Build deduplicated order material name options for mapping dropdowns
    _seen_ord: set = set()
    order_options: list = []  # [(raw, norm), ...]
    for _l in order_lines:
        _n = _l["norm_material"]
        if _n not in _seen_ord:
            _seen_ord.add(_n)
            order_options.append((_l["material"], _n))
    order_options.sort(key=lambda x: x[0])
    map_display_opts = ["— Sin mapeo —"] + [raw for raw, _ in order_options]
    map_norm_by_raw = {raw: nrm for raw, nrm in order_options}
    map_raw_by_norm = {nrm: raw for raw, nrm in order_options}

    # ── Lista A ────────────────────────────────────────────────────────────────
    st.markdown(
        '<div class="section-header">📋 LISTA A — Materiales SIN órdenes (asignación manual)</div>',
        unsafe_allow_html=True,
    )
    if list_a:
        df_a = pd.DataFrame([{
            "Material": m["name"],
            "On Hand": m["on_hand"],
            "Asset Value ($)": m["asset_value"],
            "Acción": "⚠ Asignación manual requerida",
        } for m in list_a])
        st.dataframe(df_a.style.format({"Asset Value ($)": "${:,.2f}"}),
                     use_container_width=True)
    else:
        st.success("✅ Todos los materiales con stock encontraron al menos una orden.")

    # ── Definir mapeos manuales ────────────────────────────────────────────────
    if list_a:
        st.markdown(
            '<div class="section-header">🔗 Definir mapeos manuales — materiales sin match</div>',
            unsafe_allow_html=True,
        )
        st.caption(
            "Selecciona a qué nombre en las órdenes corresponde cada material sin match. "
            "Los mapeos persisten entre sesiones (descarga el JSON desde el panel lateral)."
        )
        _cur_maps = get_project_mappings(active)
        with st.form("manual_mappings_form"):
            for _mat in list_a:
                _inv_norm = _mat["norm_name"]
                _cur_order_norm = _cur_maps.get(_inv_norm)
                _default_idx = 0
                if _cur_order_norm and _cur_order_norm in map_raw_by_norm:
                    try:
                        _default_idx = map_display_opts.index(map_raw_by_norm[_cur_order_norm])
                    except ValueError:
                        _default_idx = 0
                st.selectbox(
                    f"**{_mat['name']}**",
                    options=map_display_opts,
                    index=_default_idx,
                    key=f"map_{_inv_norm}",
                )
            _submitted_a = st.form_submit_button("💾 Guardar mapeos y re-ejecutar", type="primary")
        if _submitted_a:
            _new_maps = dict(_cur_maps)
            for _mat in list_a:
                _inv_norm = _mat["norm_name"]
                _sel = st.session_state.get(f"map_{_inv_norm}", "— Sin mapeo —")
                if _sel == "— Sin mapeo —":
                    _new_maps.pop(_inv_norm, None)
                else:
                    _new_maps[_inv_norm] = map_norm_by_raw[_sel]
            save_project_mappings(active, _new_maps)
            st.session_state.auto_run = True
            st.rerun()

    # ── Lista B ────────────────────────────────────────────────────────────────
    st.markdown(
        '<div class="section-header">⚠️ LISTA B — Matches DUDOSOS (requiere confirmación)</div>',
        unsafe_allow_html=True,
    )
    if list_b:
        df_b = pd.DataFrame([{
            "Nombre en Inv.": item["name_inv"],
            "Nombre en Pedido": item["name_order"],
            "Pedido #": item["pedido"],
            "Rubro": item["rubro"],
            "Similitud": item["score"],
        } for item in list_b])
        st.dataframe(df_b.style.format({"Similitud": "{:.1%}"}),
                     use_container_width=True)
        st.caption("Estos matches se procesaron automáticamente. Confirma los correctos o corrige los incorrectos.")

        # Deduplicate list_b by norm_inv so each inv material appears once
        _seen_b: set = set()
        _listb_dedup: list = []
        for _item in list_b:
            if _item["norm_inv"] not in _seen_b:
                _seen_b.add(_item["norm_inv"])
                _listb_dedup.append(_item)

        _cur_maps_b = get_project_mappings(active)
        _confirm_opts = ["✅ Confirmar match actual"] + map_display_opts  # includes "— Sin mapeo —"
        with st.form("listb_review_form"):
            st.caption("Confirma o corrige cada match dudoso:")
            for _item in _listb_dedup:
                st.markdown(
                    f"**Inv:** {_item['name_inv']} &nbsp;→&nbsp; **Orden:** {_item['name_order']}"
                    f" *(similitud: {_item['score']:.1%})*"
                )
                st.selectbox(
                    "Acción:",
                    options=_confirm_opts,
                    index=0,
                    key=f"listb_{_item['norm_inv']}",
                    label_visibility="collapsed",
                )
            _submitted_b = st.form_submit_button("💾 Guardar correcciones de Lista B", type="secondary")
        if _submitted_b:
            _new_maps_b = dict(_cur_maps_b)
            for _item in _listb_dedup:
                _sel = st.session_state.get(f"listb_{_item['norm_inv']}", "✅ Confirmar match actual")
                if _sel == "✅ Confirmar match actual":
                    # Save the current fuzzy match as a manual mapping
                    _new_maps_b[_item["norm_inv"]] = _item["norm_order"]
                elif _sel != "— Sin mapeo —":
                    _new_maps_b[_item["norm_inv"]] = map_norm_by_raw[_sel]
            save_project_mappings(active, _new_maps_b)
            st.session_state.auto_run = True
            st.rerun()
    else:
        st.success("✅ No hay matches dudosos — todos los nombres coincidieron exactamente.")
 
    # ── Lista C ────────────────────────────────────────────────────────────────
    st.markdown(
        '<div class="section-header">🔵 LISTA C — Materiales con líneas en estado PENDIENTE</div>',
        unsafe_allow_html=True,
    )
    if list_c:
        df_c = pd.DataFrame(list_c)
        df_c.columns = ["Material", "Pedido #", "Fecha", "Cantidad usada", "Rubro asignado", "Estado original"]
        st.dataframe(df_c, use_container_width=True)
        st.caption("Estos materiales fueron adjudicados desde órdenes Pendientes (no entregadas físicamente).")
    else:
        st.info("ℹ No se usaron líneas con estado Pendiente en la distribución.")
 
    # ── Sobrantes asignados ────────────────────────────────────────────────────
    st.markdown(
        '<div class="section-header">🟡 LISTA D — Sobrante de stock asignado al último rubro</div>',
        unsafe_allow_html=True,
    )
    if remainder_alerts:
        df_d = pd.DataFrame(remainder_alerts)
        df_d.columns = [
            "Material", "On Hand", "Cubierto por órdenes",
            "Sobrante", "Rubro asignado (automático)",
            "Pedido referencia", "Fecha pedido",
        ]
        st.dataframe(
            df_d.style.format({
                "On Hand": "{:,.4f}",
                "Cubierto por órdenes": "{:,.4f}",
                "Sobrante": "{:,.4f}",
            }).map(
                lambda v: "background-color: #fff8e1",
                subset=["Sobrante", "Rubro asignado (automático)"]
            ),
            use_container_width=True,
        )
        st.caption(
            "⚠ El stock de estos materiales supera las cantidades en órdenes. "
            "El sobrante fue asignado automáticamente al rubro del pedido más reciente. "
            "Revise si corresponde hacer una asignación manual diferente."
        )
    else:
        st.success("✅ Todos los materiales quedaron completamente cubiertos por sus órdenes.")
 
    # ── Rubros no encontrados ──────────────────────────────────────────────────
    if unmatched_rubros:
        st.markdown(
            '<div class="section-header">❗ Rubros en órdenes no encontrados en hoja Inv.</div>',
            unsafe_allow_html=True,
        )
        df_rub = pd.DataFrame(unmatched_rubros, columns=["Código", "Descripción completa"])
        st.dataframe(df_rub, use_container_width=True)
        st.caption("Las cantidades asignadas a estos rubros NO fueron escritas en el archivo (columna no existe).")
 
    # ── Orders date breakdown ──────────────────────────────────────────────────
    with st.expander("📅 Detalle de órdenes procesadas"):
        dates_info = {}
        for l in order_lines:
            key = (l["pedido_name"], l["fecha"].strftime("%d/%m/%Y"))
            dates_info[key] = dates_info.get(key, 0) + 1
        df_dates = pd.DataFrame([
            {"Pedido": k[0], "Fecha": k[1], "Líneas": v}
            for k, v in sorted(dates_info.items())
        ])
        st.dataframe(df_dates, use_container_width=True)
